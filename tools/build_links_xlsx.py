#!/usr/bin/env python3
"""Build the contractor status link workbook from published-page state."""
import json, glob, os, sys, collections
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(__file__))
from build_rep_pages import canon  # noqa: E402

ROOT = os.path.join(os.path.dirname(__file__), '..')
CFG = json.load(open(os.path.join(ROOT, 'config', 'contractors.json')))
MAN = json.load(open(os.path.join(ROOT, 'work', 'pages', 'manifest.json')))

pub = {}
for t in glob.glob(os.path.join(ROOT, 'work', 'publish', 'done*.tsv')):
    for line in open(t):
        p = line.rstrip('\n').split('\t')
        if len(p) >= 2 and p[1].strip():
            pub[p[0].strip()] = p[1].strip()
for e in glob.glob(os.path.join(ROOT, 'work', 'publish', 'results*.json')):
    for r in json.load(open(e)):
        pub.setdefault(r['file'], r['shareId'])

COMPANY = {'LINEAR roofing': 'Linear Roofing', 'Strong House Pro': 'Stronghouse Pro'}
URL = "https://www.send.co/a/{}"

reps, unassigned = [], []
for m in MAN:
    sid = pub.get(m['file']) or m.get('shareId')
    if not sid:
        continue
    row = {'company': COMPANY.get(m['label'], m['label']), 'rep': m['rep'],
           'files': m['files'], 'sid': sid}
    (unassigned if 'unassigned' in m['file'] else reps).append(row)
reps.sort(key=lambda r: (r['company'], -r['files'], r['rep']))
unassigned.sort(key=lambda r: r['company'])

live_ids = {r['sid'] for r in reps} | {u['sid'] for u in unassigned}
superseded = []
for c in CFG['contractors']:
    for s in c.get('supersededPages', []):
        match = next((r for r in reps if r['company'] == COMPANY.get(c['label'], c['label'])
                      and canon(r['rep']) == canon(s['name'])), None)
        superseded.append({'company': COMPANY.get(c['label'], c['label']), 'old': s['name'],
                           'sid': s['shareId'],
                           'now': URL.format(match['sid']) if match else 'company unassigned page'})

company_pages = [{'client': c['company'], 'sid': c['shareId']}
                 for c in CFG['contractors']
                 if c.get('pageMode') != 'per-rep' and c.get('shareId')]

# ---------------------------------------------------------------- styling
F = 'Arial'
HDR_FILL = PatternFill('solid', fgColor='1F3864')
HDR_FONT = Font(name=F, bold=True, color='FFFFFF', size=11)
WARN_FILL = PatternFill('solid', fgColor='C00000')
NOTE_FONT = Font(name=F, italic=True, size=9, color='595959')
LINK_FONT = Font(name=F, size=10, color='0563C1', underline='single')
BODY = Font(name=F, size=10)
BOLD = Font(name=F, size=10, bold=True)
THIN = Side(style='thin', color='D9D9D9')
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def header(ws, cols, row=1, fill=HDR_FILL):
    for i, (title, width) in enumerate(cols, start=1):
        c = ws.cell(row=row, column=i, value=title)
        c.fill, c.font, c.border = fill, HDR_FONT, BOX
        c.alignment = Alignment(horizontal='left', vertical='center')
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[row].height = 22
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def put(ws, r, c, v, font=BODY, link=None, num=None):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font, cell.border = font, BOX
    if link:
        cell.hyperlink, cell.font = link, LINK_FONT
    if num:
        cell.number_format = num
    return cell


wb = Workbook()

# ---------------------------------------------------------------- 1. rep links
ws = wb.active
ws.title = 'Rep Links'
header(ws, [('Company', 18), ('Rep', 28), ('Open Files', 11), ('Status Page Link', 46)])
r = 2
for row in reps:
    put(ws, r, 1, row['company'])
    put(ws, r, 2, row['rep'])
    put(ws, r, 3, row['files'], num='0')
    put(ws, r, 4, URL.format(row['sid']), link=URL.format(row['sid']))
    r += 1
for row in unassigned:
    put(ws, r, 1, row['company'], font=BOLD)
    put(ws, r, 2, 'Unassigned files (no rep on card)', font=BOLD)
    put(ws, r, 3, row['files'], font=BOLD, num='0')
    put(ws, r, 4, URL.format(row['sid']), link=URL.format(row['sid']))
    r += 1
tot = r
put(ws, tot, 2, 'TOTAL', font=BOLD)
put(ws, tot, 3, f'=SUM(C2:C{r-1})', font=BOLD, num='0')
put(ws, tot, 4, f'=COUNTA(D2:D{r-1})&" pages"', font=BOLD)
ws.cell(row=tot + 2, column=1,
        value='Send each rep only the link on their row. Counts are open files at the time of '
              'the 2026-08-15 board export.').font = NOTE_FONT

# ---------------------------------------------------------------- 2. company links
ws2 = wb.create_sheet('Company Links')
header(ws2, [('Client', 34), ('Status Page Link', 46)])
r = 2
for cp in company_pages:
    put(ws2, r, 1, cp['client'])
    put(ws2, r, 2, URL.format(cp['sid']), link=URL.format(cp['sid']))
    r += 1
put(ws2, r, 1, 'TOTAL', font=BOLD)
put(ws2, r, 2, f'=COUNTA(B2:B{r-1})&" pages"', font=BOLD)
ws2.cell(row=r + 2, column=1,
         value='One page per firm — all their reps share it.').font = NOTE_FONT

# ---------------------------------------------------------------- 3. superseded
ws3 = wb.create_sheet('DO NOT SEND')
header(ws3, [('Company', 18), ('Old Page', 24), ('Old Link (do not send)', 40),
             ('Current Page Instead', 40)], fill=WARN_FILL)
r = 2
for s in superseded:
    put(ws3, r, 1, s['company'])
    put(ws3, r, 2, s['old'])
    put(ws3, r, 3, URL.format(s['sid']))
    put(ws3, r, 4, s['now'], link=s['now'] if s['now'].startswith('http') else None)
    r += 1
put(ws3, r, 2, 'TOTAL', font=BOLD)
put(ws3, r, 3, f'=COUNTA(C2:C{r-1})&" stale links"', font=BOLD)
for i, line in enumerate([
        'These pages were replaced during the 2026-08-15 rebuild and still show an outdated file list.',
        'They were left live by choice. Column C links are stale — send the column D link instead.',
        'Cause: the original roster used initials from email addresses (M. Pierce) before the board',
        'export supplied full names (Melvin Pierce), so a second page was created rather than updated.']):
    ws3.cell(row=r + 2 + i, column=1, value=line).font = NOTE_FONT

out = os.path.join(ROOT, 'work', 'Contractor-Status-Links.xlsx')
wb.save(out)
print(f"reps={len(reps)} unassigned={len(unassigned)} company={len(company_pages)} superseded={len(superseded)}")
print(f"written: {out}")
